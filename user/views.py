from django.contrib.auth import get_user_model, logout
from django.contrib.auth.models import Group
from django.shortcuts import redirect
from django.urls import reverse_lazy
import requests
import environ
env = environ.Env(DEBUG=(bool, False))

from django.views.generic import (
    CreateView,
    DetailView,
    ListView,
    UpdateView,
    View,
)
from root.utils import DeleteMixin, remove_from_DB
from user.permission import IsAdminMixin, AdminBillingMixin

from .forms import UserCreateForm, UserForm, AdminForm


User = get_user_model()


class UserMixin(IsAdminMixin):
    model = User
    form_class = UserCreateForm
    paginate_by = 50
    queryset = User.objects.filter(status=True)
    success_url = reverse_lazy("user:user_list")
    search_lookup_fields = ["username", "email", "full_name"]


class UserList(UserMixin, ListView):
    template_name = "user/user_list.html"
    queryset = User.objects.filter(status=True, is_deleted=False, groups__name__in=["admin"])

    def get_queryset(self, *args, **kwargs):
        queryset = self.queryset.exclude(id=self.request.user.id)
        return queryset


class UserDetail(UserMixin, DetailView):
    template_name = "user/user_detail.html"


class UserCreate(UserMixin, CreateView):
    template_name = "create.html"

    def form_valid(self, form):
        form.instance.is_superuser = False
        form.instance.is_staff = True
        form.instance.organization = self.request.user.organization
        object = form.save()
        # FLASK_URL = env("FLASK_USER_CREATE_URL")
        # TOKEN = env("FLASK_USER_CREATE_KEY")
        # data= {
        #     "token":TOKEN,
        #     "username": object.username,
        #     "baseURL":self.request.scheme+'://'+self.request.get_host()
        # }
        # requests.post(
        #     FLASK_URL,
        #     json=data
        # )
        group, _ = Group.objects.get_or_create(name='admin')
        object.groups.add(group)
        return super().form_valid(form)
        


class UserAdmin(UserMixin, CreateView):
    template_name = "create.html"


class UserUpdate(UserMixin, UpdateView):
    template_name = "update.html"

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        old_username = self.object.username
        p = super().post(request, *args, **kwargs)
        # FLASK_URL = env("FLASK_USER_UPDATE_URL")
        # TOKEN = env("FLASK_USER_CREATE_KEY")
        # data= {
        #     "token":TOKEN,
        #     "username":old_username,
        #     "newUsername": request.POST.get('username'),
        #     "baseURL":request.scheme+'://'+request.get_host(),
        #     "type":"UPDATE"
        # }
        # response = requests.post(
        #     FLASK_URL,
        #     json=data
        # )
        return p
        


class UserDelete(UserMixin, View):
    def get(self, request):
        status = remove_from_DB(self, request)
        return JsonResponse({"deleted": status})


def logout_user(request):
    logout(request)
    return redirect(reverse_lazy("user:login_view"))


from django.db import transaction
from django.http import JsonResponse
from django.urls import reverse_lazy
from django.views.generic import (
    CreateView,
    DetailView,
    ListView,
    TemplateView,
    UpdateView,
    View,
)
from root.utils import DeleteMixin
from .models import Customer
from .forms import CustomerForm


class CustomerMixin(AdminBillingMixin):
    model = Customer
    form_class = CustomerForm
    paginate_by = 50
    queryset = Customer.objects.filter(status=True, is_deleted=False)
    success_url = reverse_lazy("user:customer_list")
    search_lookup_fields = ["name", "tax_number", "contact_number", "email"]


class CustomerList(CustomerMixin, ListView):
    template_name = "customer/customer_list.html"
    queryset = Customer.objects.active()


class CustomerDetail(CustomerMixin, DetailView):
    template_name = "customer/customer_detail.html"


class CustomerCreate(CustomerMixin, CreateView):
    template_name = "create.html"


class CustomerUpdate(CustomerMixin, UpdateView):
    template_name = "update.html"


class CustomerDelete(CustomerMixin, DeleteMixin, View):
    pass


class AgentMixin(IsAdminMixin):
    model = User
    form_class = UserForm
    paginate_by = 50
    queryset = User.objects.filter(status=True, is_deleted=False)
    success_url = reverse_lazy("user:agent_list")
    search_lookup_fields = ["username", "email", "full_name"]


class AgentList(AgentMixin, ListView):
    template_name = "agent/agent_list.html"
    queryset = User.objects.filter(
        groups__name__in=["agent"], status=True, is_deleted=False
    )


class AgentCreate(AgentMixin, CreateView):
    template_name = "create.html"

    def form_valid(self, form):

        form.instance.is_superuser = False
        form.instance.is_staff = True
        object = form.save()

        group, created = Group.objects.get_or_create(name="agent")
        object.groups.add(group)
        return super().form_valid(form)


class AgentUpdate(AgentMixin, UpdateView):
    template_name = "update.html"


class AgentDelete(AgentMixin, DeleteMixin, View):
    pass

from openpyxl import load_workbook
from .models import Customer

# from django.shortcuts import render, redirect
# from django.contrib import messages
# from openpyxl import load_workbook
# from .models import Customer

# def upload_customer_excel(request):
#     if request.method == 'POST' and request.FILES.get('file'):
#         excel_file = request.FILES['file']
#         try:
#             wb = load_workbook(excel_file)
#             sheet = wb.active  # You can also use wb.worksheets[0] if needed

#             customers_created = 0
#             customer_create_errors = []

#             for index, row in enumerate(sheet.iter_rows(min_row=2), start=2):  # Skip header row
#                 try:
#                     fname = str(row[0].value).strip() if row[0].value else ""
#                     mname = str(row[1].value).strip() if row[1].value else ""
#                     lname = str(row[2].value).strip() if row[2].value else ""
#                     roll_no = str(row[3].value).strip() if row[3].value else None
#                     student_class = str(row[4].value).strip() if row[4].value else None
#                     section = str(row[5].value).strip() if row[5].value else ""
#                     lunch_type = str(row[6].value).strip().lower() if row[6].value else ""

#                     full_name = f"{fname} {mname} {lname}".strip()

#                     if lunch_type in ['non veg', 'nonveg']:
#                         meal_preference = 'nonveg'
#                     elif lunch_type == 'veg':
#                         meal_preference = 'veg'
#                     elif lunch_type == 'egg':
#                         meal_preference = 'egg'
#                     else:
#                         meal_preference = 'unknown'

#                     Customer.objects.create(
#                         name=full_name,
#                         section=section,
#                         roll_no=roll_no,
#                         student_class=student_class,
#                         meal_preference=meal_preference
#                     )
#                     customers_created += 1

#                 except Exception as e:
#                     customer_create_errors.append(f"Row {index}: {str(e)}")

#             # Show messages
#             if customer_create_errors:
#                 messages.error(request, "Some entries could not be imported:\n" + "\n".join(customer_create_errors), extra_tags='danger')

#             if customers_created:
#                 messages.success(request, f"{customers_created} students imported successfully.", extra_tags='success')

#         except Exception as e:
#             messages.error(request, f"Error reading file: {e}", extra_tags='danger')

#         return redirect('user:customer_list')

#     return redirect('user:customer_list')

from django.shortcuts import render, redirect
from django.contrib import messages
from openpyxl import load_workbook
from .models import Customer

def upload_customer_excel(request):
    if request.method == 'POST' and request.FILES.get('file'):
        excel_file = request.FILES['file']
        try:
            wb = load_workbook(excel_file)
            sheet = wb.active  # You can also use wb.worksheets[0] if needed

            customers_created = 0
            customer_create_errors = []

            for index, row in enumerate(sheet.iter_rows(min_row=2), start=2):  # Skip header row
                try:
                    # fname = str(row[0].value).strip() if row[0].value else ""
                    # mname = str(row[1].value).strip() if row[1].value else ""
                    # lname = str(row[2].value).strip() if row[2].value else ""
                    # roll_no = str(row[3].value).strip() if row[3].value else None
                    # student_class = str(row[4].value).strip() if row[4].value else None
                    # section = str(row[5].value).strip() if row[5].value else ""
                    # lunch_type = str(row[6].value).strip().lower() if row[6].value else ""

                    # fname = str(row[0].value).strip() if row[0].value else ""
                    # # mname = str(row[1].value).strip() if row[1].value else ""
                    # lname = str(row[1].value).strip() if row[1].value else ""
                    # roll_no = str(row[2].value).strip() if row[2].value else None
                    # student_class = str(row[3].value).strip() if row[3].value else None
                    # section = str(row[5].value).strip() if row[5].value else ""
                    # lunch_type = str(row[4].value).strip().lower() if row[4].value else ""

                    # # full_name = f"{fname} {mname} {lname}".strip()
                    # full_name = f"{fname} {lname}".strip()

                    full_name = str(row[0].value).strip() if row[0].value else ""
                    roll_no = str(row[1].value).strip() if row[1].value else None                    
                    lunch_type = str(row[2].value).strip().lower() if row[2].value else ""
                    section = str(row[3].value).strip() if row[3].value else ""
                    student_class = int(row[4].value) if row[4].value else None


                    if lunch_type in ['non veg', 'nonveg', 'non-veg']:
                        meal_preference = 'nonveg'
                    elif lunch_type in ['veg', 'Veg']:
                        meal_preference = 'veg'
                    elif lunch_type in ['egg', 'Egg']:
                        meal_preference = 'egg'
                    else:
                        meal_preference = 'unknown'

                    Customer.objects.create(
                        name=full_name,
                        section=section,
                        roll_no=roll_no,
                        student_class=student_class,
                        meal_preference=meal_preference
                    )
                    customers_created += 1

                except Exception as e:
                    customer_create_errors.append(f"Row {index}: {str(e)}")

            # Show messages
            if customer_create_errors:
                messages.error(request, "Some entries could not be imported:\n" + "\n".join(customer_create_errors), extra_tags='danger')

            if customers_created:
                messages.success(request, f"{customers_created} students imported successfully.", extra_tags='success')

        except Exception as e:
            messages.error(request, f"Error reading file: {e}", extra_tags='danger')

        return redirect('user:customer_list')

    return redirect('user:customer_list')
